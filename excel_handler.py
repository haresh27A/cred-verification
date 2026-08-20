"""
excel_handler.py
----------------
Read input Excel/CSV files (.csv, .xlsx, .xls) and save/export styled Excel output.
Preserves original columns and data while adding highlights, cell comments, and notes.
Credential and NPI Finder
"""

import os
import io
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

from utils import ensure_dir


def read_input_file(file_source, filename=None):
    """
    Read CSV or Excel (.xlsx, .xls) file from file path or file-like bytes stream.
    Supports delimiter auto-detection and preserves all original data rows/columns.
    """
    try:
        if not isinstance(file_source, (str, io.StringIO, io.BytesIO)):
            content = file_source.read()
            if isinstance(content, str):
                file_source = io.StringIO(content)
            else:
                file_source = io.BytesIO(content)

        name = (filename or (file_source if isinstance(file_source, str) else "")).lower()

        if name.endswith((".xlsx", ".xls")):
            try:
                df = pd.read_excel(file_source, dtype=str).fillna("")
            except Exception:
                if hasattr(file_source, "seek"):
                    file_source.seek(0)
                df = pd.read_excel(file_source, engine="openpyxl", dtype=str).fillna("")
        else:  # Default CSV / TSV text format
            try:
                df = pd.read_csv(file_source, sep=None, engine="python", dtype=str, keep_default_na=False)
            except Exception:
                if hasattr(file_source, "seek"):
                    file_source.seek(0)
                df = pd.read_csv(file_source, sep="\t", dtype=str, keep_default_na=False)
                if len(df.columns) <= 1:
                    if hasattr(file_source, "seek"):
                        file_source.seek(0)
                    df = pd.read_csv(file_source, sep=",", dtype=str, keep_default_na=False)

        # Standardize column headers without destroying original names
        df.columns = [str(column).strip() for column in df.columns]
        return df

    except Exception as e:
        raise Exception(f"Unable to read uploaded file: {e}")


def prepare_output_dataframe(df):
    """
    Ensure standard result columns exist without overwriting existing input data.
    """
    target_columns = [
        "NPI 1",
        "NPI1",
        "Credential",
        "Credental",
        "NPI Status",
        "NPI Entity Type",
        "NPI Validation",
        "Match Confidence",
        "Validation Notes",
        "Url",
        "NPI 1 Url",
        "NPI1 Url"
    ]
    for column in target_columns:
        if column not in df.columns:
            df[column] = ""

    return df


def apply_professional_excel_styling(workbook, invalid_npi_rows=None, mismatch_rows=None, npi_notes_map=None):
    """
    Apply openpyxl styling:
    - Dark Navy Header fill (#1E293B) with white bold text.
    - Light-red fill (#FCE8E6) with dark red bold text (#991B1B) for invalid/mismatched NPI cells.
    - Add openpyxl Cell Comments on invalid/mismatched NPI cells explaining the problem.
    - Clickable hyperlinks for Url and NPI 1 Url columns.
    - Gridlines enabled and dynamic column widths.
    """
    invalid_npi_rows = invalid_npi_rows or []
    mismatch_rows = mismatch_rows or []
    npi_notes_map = npi_notes_map or {}
    all_highlight_rows = list(set(invalid_npi_rows + mismatch_rows))

    worksheet = workbook.active
    worksheet.views.sheetView[0].showGridLines = True

    column_map = {}
    for cell in worksheet[1]:
        if cell.value is not None:
            column_map[str(cell.value).strip()] = cell.column

    # Header styling
    header_fill = PatternFill(fill_type="solid", fgColor="1E293B")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB")
    )

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border

    worksheet.row_dimensions[1].height = 28

    # Red warning highlight style
    red_fill = PatternFill(fill_type="solid", fgColor="FCE8E6")
    red_font = Font(name="Calibri", size=10, color="991B1B", bold=True)

    # Standard cell fonts
    cell_font = Font(name="Calibri", size=10)
    link_font = Font(name="Calibri", size=10, color="1D4ED8", underline="single")

    # Format data rows
    for row_idx in range(2, worksheet.max_row + 1):
        worksheet.row_dimensions[row_idx].height = 20
        for col_idx in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.font = cell_font
            cell.border = thin_border

    # Highlight invalid or mismatched NPI cells & attach Cell Comments
    if "NPI" in column_map:
        npi_col = column_map["NPI"]
        for row_idx in all_highlight_rows:
            excel_row = row_idx + 2  # 1-indexed openpyxl header is row 1
            if excel_row <= worksheet.max_row:
                cell = worksheet.cell(row=excel_row, column=npi_col)
                cell.fill = red_fill
                cell.font = red_font

                # Attach cell comment note if available
                note_text = npi_notes_map.get(row_idx, "Invalid NPI or Entity Type Mismatch")
                cell.comment = Comment(text=note_text, author="Credential and NPI Finder")

    # Add hyperlinks for URL columns
    for col_name in ["Url", "NPI 1 Url", "NPI1 Url", "Pending CPID URL", "Suggested Admin URL"]:
        if col_name in column_map:
            col_num = column_map[col_name]
            for row_idx in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row=row_idx, column=col_num)
                val = str(cell.value or "").strip()
                if val and val.startswith("http"):
                    cell.hyperlink = val
                    cell.font = link_font

    # Adjust column widths dynamically
    for col_name, col_num in column_map.items():
        max_len = max(
            len(str(worksheet.cell(row=r, column=col_num).value or ''))
            for r in range(1, worksheet.max_row + 1)
        )
        col_letter = get_column_letter(col_num)
        worksheet.column_dimensions[col_letter].width = min(max(max_len + 3, 14), 65)

    return workbook


def save_output_file(
    df,
    output_path,
    invalid_npi_rows=None,
    mismatch_rows=None,
    npi_notes_map=None
):
    """
    Save DataFrame to Excel file on disk with openpyxl styling and cell comments.
    """
    output_folder = os.path.dirname(output_path)
    if output_folder:
        ensure_dir(output_folder)

    try:
        df.to_excel(output_path, index=False)
        workbook = load_workbook(output_path)
        workbook = apply_professional_excel_styling(
            workbook,
            invalid_npi_rows=invalid_npi_rows,
            mismatch_rows=mismatch_rows,
            npi_notes_map=npi_notes_map
        )
        workbook.save(output_path)
    except Exception as e:
        raise Exception(f"Unable to save output file: {e}")


def generate_excel_bytes(
    df,
    invalid_npi_rows=None,
    mismatch_rows=None,
    npi_notes_map=None
):
    """
    Generate styled Excel binary buffer in memory for web file downloads.
    """
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False)

    buffer.seek(0)
    workbook = load_workbook(buffer)
    workbook = apply_professional_excel_styling(
        workbook,
        invalid_npi_rows=invalid_npi_rows,
        mismatch_rows=mismatch_rows,
        npi_notes_map=npi_notes_map
    )

    out_buffer = io.BytesIO()
    workbook.save(out_buffer)
    out_buffer.seek(0)
    return out_buffer.getvalue()